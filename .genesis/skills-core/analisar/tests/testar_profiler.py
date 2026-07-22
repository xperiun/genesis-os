# -*- coding: utf-8 -*-
"""
Regressão do profiler. Cada caso aqui já produziu número errado SEM AVISAR em
teste real, então cada um vira asserção: o profiler tem que detectar, e tem que
NÃO detectar o que não é problema (alarme falso ensina a ignorar o alarme).

Rode:  python testar_profiler.py
Saída: uma linha por caso, e código 1 se qualquer um falhar.
"""

import os
import subprocess
import sys

AQUI = os.path.dirname(os.path.abspath(__file__))
FIXTURES = os.path.join(AQUI, "fixtures")
PROFILER = os.path.join(AQUI, "..", "scripts", "perfilar.py")

sys.path.insert(0, os.path.join(AQUI, "..", "scripts"))
import perfilar as P  # noqa: E402

falhas = []


def _tipos(rel, severidade=None):
    return {a["tipo"] for a in rel["armadilhas"]
            if severidade is None or a["severidade"] == severidade}


def checar(rotulo, condicao, detalhe=""):
    print(f"  {'ok  ' if condicao else 'FALHA'}  {rotulo}" + (f"  ({detalhe})" if detalhe and not condicao else ""))
    if not condicao:
        falhas.append(rotulo)


def caso_vendas_alvorada():
    print("\nvendas_alvorada.xlsx  (miolo institucional + subtotais + texto BR)")
    r = P.perfilar(os.path.join(FIXTURES, "vendas_alvorada.xlsx"))
    checar("acha o cabeçalho na linha 7", r["cabecalho"]["linha"] == 7, f"achou {r['cabecalho']['linha']}")
    checar("conta 7 linhas de dado", r["contagem"]["linhas_de_dado"] == 7,
           f"contou {r['contagem']['linhas_de_dado']}")
    checar("acha as 4 linhas de agregação", r["contagem"]["linhas_de_agregacao"] == 4,
           f"achou {r['contagem']['linhas_de_agregacao']}")
    checar("marca agregação como P0", "agregacao" in _tipos(r, "P0"))
    checar("NÃO marca mescla decorativa do topo", "merge" not in _tipos(r),
           f"marcou {r['merges_total']} mescla(s)")
    checar("pula a aba de instruções", r["aba_escolhida"]["nome"] == "Vendas 2025",
           f"escolheu {r['aba_escolhida']['nome']}")
    fat = next(c for c in r["colunas"] if c["nome"] == "Faturamento")
    checar("prova convenção BR no faturamento", fat["numerico"]["convencao"] == "BR",
           fat["numerico"]["convencao"])
    dat = next(c for c in r["colunas"] if c["nome"] == "Data")
    checar("prova data BR (dia > 12 existe)", dat["data"]["veredito"] == "BR", dat["data"]["veredito"])


def caso_coluna_mista():
    print("\ncoluna_mista.xlsx  (o erro de 37x)")
    r = P.perfilar(os.path.join(FIXTURES, "coluna_mista.xlsx"))
    checar("detecta tipo misto como P0", "tipo_misto" in _tipos(r, "P0"))
    col = next(c for c in r["colunas"] if c["nome"] == "Valor")
    checar("conta os dois tipos na mesma coluna",
           col["numerico"]["numeros_nativos"] == 3 and col["numerico"]["texto_parseavel"] == 3,
           f"nativos={col['numerico']['numeros_nativos']} texto={col['numerico']['texto_parseavel']}")


def caso_ambiguo():
    print("\nambiguo.xlsx  (nada pode ser provado, a skill tem que PERGUNTAR)")
    r = P.perfilar(os.path.join(FIXTURES, "ambiguo.xlsx"))
    checar("marca separador ambíguo como P0", "separador_ambiguo" in _tipos(r, "P0"))
    checar("marca data ambígua como P0", "data_ambigua" in _tipos(r, "P0"))
    dat = next(c for c in r["colunas"] if c["nome"] == "Data")
    checar("não finge que provou a data", dat["data"]["veredito"] == "AMBIGUO", dat["data"]["veredito"])


def caso_merge():
    print("\nmerge_nos_dados.xlsx  (some 60% do dinheiro no agrupamento)")
    r = P.perfilar(os.path.join(FIXTURES, "merge_nos_dados.xlsx"))
    checar("detecta mescla dentro dos dados como P0", "merge" in _tipos(r, "P0"))
    checar("conta as 2 mesclas relevantes", r["merges_total"] == 2, f"contou {r['merges_total']}")


def caso_csv_br():
    print("\nexport_br.csv  (cp1252 + ponto e vírgula)")
    r = P.perfilar(os.path.join(FIXTURES, "export_br.csv"))
    checar("detecta o encoding legado", r["arquivo"]["encoding"] in ("cp1252", "latin-1"),
           str(r["arquivo"]["encoding"]))
    checar("detecta o delimitador ponto e vírgula", r["arquivo"]["delimitador"] == ";",
           repr(r["arquivo"]["delimitador"]))
    checar("conta as 3 linhas", r["contagem"]["linhas_de_dado"] == 3,
           f"contou {r['contagem']['linhas_de_dado']}")


def caso_pii():
    print("\ncom_pii.xlsx  (dado pessoal na tela é risco jurídico, não bug)")
    r = P.perfilar(os.path.join(FIXTURES, "com_pii.xlsx"))
    suspeitas = {c["nome"] for c in r["colunas"] if c["possivel_pii"]}
    checar("sinaliza CPF, Email e Salario", {"CPF", "Email", "Salario"} <= suspeitas, str(suspeitas))
    checar("NÃO sinaliza a coluna de valor comercial", "Valor Compra" not in suspeitas)


def caso_relatorio_e_fixo():
    print("\ntamanho do relatório  (tem que ser fixo, não crescer com o arquivo)")
    import openpyxl
    grande = os.path.join(FIXTURES, "_grande.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Regiao", "Valor"])
    for i in range(5000):
        ws.append([f"R{i % 5}", 100 + i])
        if i % 1000 == 999:
            ws.append([f"SUBTOTAL R{i % 5}", 999999])
    wb.save(grande)
    r = P.perfilar(grande)
    texto = P.renderizar(r)
    checar("leu todas as linhas", r["contagem"]["linhas_de_dado"] == 5000,
           f"leu {r['contagem']['linhas_de_dado']}")
    checar("achou os 5 subtotais no corpo", r["contagem"]["linhas_de_agregacao"] == 5,
           f"achou {r['contagem']['linhas_de_agregacao']}")
    checar("relatório continua curto (< 6 KB)", len(texto) < 6000, f"{len(texto)} chars")
    os.remove(grande)


def caso_recusa_binario():
    print("\nformato não suportado  (tem que falhar explicando, não com traceback)")
    alvo = os.path.join(FIXTURES, "_x.docx")
    open(alvo, "wb").write(b"PK\x03\x04lixo")
    p = subprocess.run([sys.executable, PROFILER, alvo], capture_output=True, text=True)
    saida = (p.stdout + p.stderr).lower()
    checar("explica em português o que fazer", "não sei ler" in saida or "nao sei ler" in saida)
    checar("não vaza traceback", "traceback" not in saida)
    os.remove(alvo)


def caso_reconciliacao():
    print("\nreconciliação  (o passe 3: nenhum número sem certidão)")
    import openpyxl
    from reconciliar import Auditoria, parse_numero, marcadores_de_agregacao, mascarar

    wb = openpyxl.load_workbook(os.path.join(FIXTURES, "vendas_alvorada.xlsx"), data_only=True)
    linhas = [list(r) for r in wb["Vendas 2025"].iter_rows(min_row=8, values_only=True)]
    linhas = [l for l in linhas if any(c is not None for c in l)]

    ingenuo = round(sum(parse_numero(l[4], "BR") or 0 for l in linhas), 2)
    checar("o erro ingênuo ainda existe (3x)", abs(ingenuo - 294862.80) < 0.01, str(ingenuo))

    limpas, agreg = marcadores_de_agregacao(linhas)
    a = Auditoria("Faturamento total de 2025")
    a.linhas_lidas = len(linhas)
    a.excluir(len(agreg), "linhas de TOTAL/SUBTOTAL")
    total = a.soma(a.numeros([l[4] for l in limpas], "BR", "a coluna Faturamento"))
    ref = max(parse_numero(l[4], "BR") or 0 for l in agreg)
    bateu = a.conferir(total, ref, "a linha TOTAL GERAL da planilha")

    checar("o número auditado é o correto", abs(total - 98287.60) < 0.01, str(total))
    checar("reconcilia contra o TOTAL GERAL do arquivo", bateu)
    checar("libera a exibição quando confere", not a.bloqueado())
    checar("a certidão declara o denominador", "7 linha(s) entraram" in a.certidao(total))

    ruim = Auditoria("teste")
    ruim.linhas_lidas = 2
    t2 = ruim.soma(ruim.numeros(["100,00", "200,00"], "BR"))
    ruim.conferir(t2, 999.99, "referência de teste")
    checar("BLOQUEIA quando não confere", ruim.bloqueado())
    # O número calculado APARECE de propósito (sem ele não dá pra investigar),
    # mas rotulado como inutilizável e sem o cabeçalho de resultado liberado.
    cert = ruim.certidao(t2)
    checar("marca o número como não utilizável", "NÃO USE ESTE NÚMERO" in cert)
    checar("não apresenta o bloqueado como resposta", "NÃO LIBERADO" in cert
           and "entraram no cálculo" not in cert)

    # o float nativo não pode ser destruído pela limpeza de texto BR
    checar("preserva float nativo (o erro de 37x)", parse_numero(12450.80, "BR") == 12450.80,
           str(parse_numero(12450.80, "BR")))
    checar("converte texto BR no mesmo passe", parse_numero("12.450,80", "BR") == 12450.80)
    checar("recusa 1.234 ambíguo em AUTO", parse_numero("1.234", "AUTO") is None)

    checar("mascara telefone sem sobrar parêntese",
           mascarar("(11) 98765-4321") == "(**) *****-****", mascarar("(11) 98765-4321"))
    checar("NÃO mascara valor monetário", mascarar("98.287,60") == "98.287,60")


if __name__ == "__main__":
    if not os.path.isdir(FIXTURES):
        print("fixtures ausentes. Rode antes: python gerar_fixtures.py")
        sys.exit(1)
    for caso in (caso_vendas_alvorada, caso_coluna_mista, caso_ambiguo, caso_merge,
                 caso_csv_br, caso_pii, caso_relatorio_e_fixo, caso_recusa_binario,
                 caso_reconciliacao):
        caso()
    print("\n" + "=" * 60)
    if falhas:
        print(f"{len(falhas)} FALHA(S): " + "; ".join(falhas))
        sys.exit(1)
    print("todos os casos passaram.")
