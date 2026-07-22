# -*- coding: utf-8 -*-
"""
Gera as planilhas hostis que o profiler tem que sobreviver.

Nada aqui é exótico. É o que sai de ERP e de Excel de PME brasileira todo dia:
logo e razão social ocupando as 6 primeiras linhas, subtotal por região no meio
dos dados, número exportado como texto em formato BR, e a coluna que tem metade
digitada à mão (float de verdade) e metade vinda do sistema (texto).

Rode: python gerar_fixtures.py
Os arquivos caem em ./fixtures/ e são consumidos pelo testar_profiler.py.
"""

import os
import openpyxl
from openpyxl.styles import Font

AQUI = os.path.dirname(os.path.abspath(__file__))
DESTINO = os.path.join(AQUI, "fixtures")


def _salvar(wb, nome):
    os.makedirs(DESTINO, exist_ok=True)
    caminho = os.path.join(DESTINO, nome)
    wb.save(caminho)
    print(f"  gerado: {nome}")
    return caminho


def vendas_alvorada():
    """O caso completo: miolo institucional, subtotais, texto BR, aba distratora."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas 2025"

    ws["A1"] = "COMERCIAL ALVORADA LTDA"
    ws.merge_cells("A1:F1")                      # mescla DECORATIVA, não pode virar P0
    ws["A2"] = "Relatório de Vendas · Exercício 2025"
    ws.merge_cells("A2:F2")
    ws["A4"] = "Emitido em:"
    ws["B4"] = "15/01/2026"
    ws["A5"] = "Responsável:"
    ws["B5"] = "Depto Financeiro"

    for i, h in enumerate(["Data", "Regiao", "Vendedor", "Produto", "Faturamento", "Margem %"], 1):
        ws.cell(row=7, column=i, value=h).font = Font(bold=True)

    linhas = [
        ("05/01/2025", "Sudeste", "Ana", "Produto A", "12.450,80", "22,5"),
        ("12/01/2025", "Sudeste", "Bruno", "Produto B", "8.300,00", "18,0"),
        ("19/02/2025", "Sudeste", "Ana", "Produto A", "15.720,45", "24,1"),
        (None, None, "TOTAL SUDESTE", None, "36.471,25", None),
        (None, None, None, None, None, None),
        ("03/03/2025", "Nordeste", "Carla", "Produto C", "22.100,00", "31,2"),
        ("11/04/2025", "Nordeste", "Diego", "Produto A", "9.870,30", "15,8"),
        (None, None, "TOTAL NORDESTE", None, "31.970,30", None),
        (None, None, None, None, None, None),
        ("07/05/2025", "Sul", "Elisa", "Produto B", "18.640,90", "27,4"),
        ("21/06/2025", "Sul", "Ana", "Produto C", "11.205,15", "19,9"),
        (None, None, "TOTAL SUL", None, "29.846,05", None),
        (None, None, None, None, None, None),
        (None, None, "TOTAL GERAL", None, "98.287,60", None),
    ]
    for r, linha in enumerate(linhas, start=8):
        for c, v in enumerate(linha, start=1):
            ws.cell(row=r, column=c, value=v)

    m = wb.create_sheet("Metas")          # distrator plausível: responder meta achando que é venda
    m["A1"], m["B1"] = "Regiao", "Meta Anual"
    for i, (reg, meta) in enumerate([("Sudeste", "40.000,00"), ("Nordeste", "35.000,00"),
                                     ("Sul", "30.000,00")], start=2):
        m.cell(row=i, column=1, value=reg)
        m.cell(row=i, column=2, value=meta)

    inst = wb.create_sheet("Instrucoes", 0)   # vira a PRIMEIRA aba, e é lixo
    inst["A1"] = "Preencher somente as células em amarelo."

    # A verdade, pro teste conferir: só as 7 linhas de transação.
    vals = [12450.80, 8300.00, 15720.45, 22100.00, 9870.30, 18640.90, 11205.15]
    return _salvar(wb, "vendas_alvorada.xlsx"), {"soma": round(sum(vals), 2), "linhas": len(vals)}


def coluna_mista():
    """
    O erro de 37x. Metade da coluna veio do sistema como texto BR, metade foi
    digitada à mão e é float de verdade. A limpeza padrão conserta o texto e
    destrói o float: 12450.8 vira 124508.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lancamentos"
    ws.append(["Descricao", "Valor"])
    for d, v in [("do sistema", "12.450,80"), ("digitado", 8300.00),
                 ("do sistema", "15.720,45"), ("digitado", 22100.5),
                 ("do sistema", "9.870,30"), ("digitado", 18640.9)]:
        ws.append([d, v])
    vals = [12450.80, 8300.00, 15720.45, 22100.5, 9870.30, 18640.9]
    return _salvar(wb, "coluna_mista.xlsx"), {"soma": round(sum(vals), 2), "linhas": len(vals)}


def ambiguo():
    """
    Nada aqui pode ser provado. 1.234 pode ser milhar ou decimal, e nenhuma data
    passa de 12 em nenhuma posição. A resposta certa da skill é PERGUNTAR.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.append(["Data", "Quantidade"])
    for d, q in [("05/01/2025", "1.234"), ("03/02/2025", "2.500"),
                 ("11/04/2025", "980"), ("07/05/2025", "1.100")]:
        ws.append([d, q])
    return _salvar(wb, "ambiguo.xlsx"), {}


def merge_nos_dados():
    """Região mesclada verticalmente. O pandas apaga e o agrupamento perde 60%."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas"
    ws.append(["Regiao", "Valor"])
    for r, (reg, val) in enumerate([("Sudeste", 100), (None, 200), (None, 150),
                                    ("Sul", 300), (None, 250)], start=2):
        ws.cell(row=r, column=1, value=reg)
        ws.cell(row=r, column=2, value=val)
    ws.merge_cells("A2:A4")
    ws.merge_cells("A5:A6")
    return _salvar(wb, "merge_nos_dados.xlsx"), {"soma": 1000, "linhas": 5}


def csv_br():
    """Export de Excel BR: cp1252, ponto e vírgula, acento."""
    os.makedirs(DESTINO, exist_ok=True)
    caminho = os.path.join(DESTINO, "export_br.csv")
    conteudo = (
        "Data;Regiao;Descricao;Faturamento\r\n"
        "05/01/2025;Sudeste;Manutenção предvista;12.450,80\r\n"
        "19/02/2025;Nordeste;Serviço avulso;8.300,00\r\n"
        "21/06/2025;Sul;Inspeção anual;15.720,45\r\n"
    ).replace("предvista", "prevista")
    with open(caminho, "wb") as f:
        f.write(conteudo.encode("cp1252"))
    print("  gerado: export_br.csv")
    return caminho, {"soma": 36471.25, "linhas": 3}


def pii():
    """Coluna com dado pessoal. Tem que ser sinalizada antes de qualquer render."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["Nome", "CPF", "Email", "Salario", "Valor Compra"])
    for n, c, e, s, v in [("Ana Souza", "123.456.789-00", "ana@exemplo.com", 4500, "1.200,00"),
                          ("Bruno Lima", "987.654.321-00", "bruno@exemplo.com", 6200, "980,50")]:
        ws.append([n, c, e, s, v])
    return _salvar(wb, "com_pii.xlsx"), {}


if __name__ == "__main__":
    print("Gerando fixtures hostis em ./fixtures/")
    for fn in (vendas_alvorada, coluna_mista, ambiguo, merge_nos_dados, csv_br, pii):
        fn()
    print("pronto.")
